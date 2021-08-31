import openpyxl

path = r"C:\Users\Fahim Mahmud\PycharmProjects\SeleniumProject\data\FoodData.xlsx"

openpyxl.load_workbook(path)
workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name("Sheet1")

# Number of rows and column
rows = sheet.max_row
cols = sheet.max_column

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value, end = "      ")
    print()


print(rows)
print(cols)
